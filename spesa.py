#!/usr/bin/env python
#spesa.py - Updates daily expenses and sums the total reached this months adn wraps it up

import openpyxl,sys,datetime
#wb = openpyxl.Workbook()
#wb.save('newspesa.xlsx')

MONTHS = ['JANUARY','FEBRUARY','MARCH','APRIL','MAY','JUNE','JULY','AUGUST','SEPTEMBER','OCTOBER','NOVEMBER','DECEMBER']

wb = openpyxl.load_workbook('/home/luho/Programming/spreadsheets/newspesa.xlsx')
path = '/home/luho/Programming/spreadsheets/newspesa.xlsx'
sheet = wb.active
new_import0 = ''.join(sys.argv[1:])
new_import = int(new_import0)
new_max = sheet.max_row
print(new_max)
sheet.cell(new_max + 1, 2).value = new_import
sheet.cell(new_max + 1, 1).value = datetime.date.today()
sheet.cell(2,3).value = '=SUM(B3:B30)'
sheet.cell(new_max+1,4).value = datetime.datetime.now().month
if datetime.datetime.now().month != sheet.cell(new_max+1,4).value:
    this_month = MONTHS[datetime.datetime.now().month -1 ]
    print(this_month)
    wb.save(str(this_month) +'.xlsx')
    sheet.delete_rows(3, 40)
wb.save(path)